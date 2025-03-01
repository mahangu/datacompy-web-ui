import pytest
import pandas as pd
import json
import numpy as np
from io import BytesIO
import openpyxl
from datacompy_web_ui.core.file_handlers import (
    FileHandler,
    CSVHandler,
    ExcelHandler,
    JSONHandler,
    ParquetHandler,
    get_handler,
)


class MockFileName:
    def __init__(self, name):
        self.name = name


def test_base_handler():
    handler = FileHandler()
    with pytest.raises(NotImplementedError):
        handler.read_data(None)
    assert handler.get_options(None) == {}


def test_csv_handler():
    handler = CSVHandler()
    # Test file type detection
    assert handler.can_handle(MockFileName("test.csv"))
    assert not handler.can_handle(MockFileName("test.xlsx"))
    # Test reading CSV data
    csv_data = "a,b\n1,2\n3,4"
    file = BytesIO(csv_data.encode())
    file.name = "test.csv"
    df = handler.read_data(file)
    assert isinstance(df, pd.DataFrame)
    assert df.shape == (2, 2)
    # Test options
    assert handler.get_options(None) == {}


def test_excel_handler():
    handler = ExcelHandler()
    # Test file type detection
    assert handler.can_handle(MockFileName("test.xlsx"))
    assert handler.can_handle(MockFileName("test.xls"))
    assert not handler.can_handle(MockFileName("test.csv"))
    # Create test Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "a"
    ws["B1"] = "b"
    ws["A2"] = 1
    ws["B2"] = 2
    # Add second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "x"
    ws2["B1"] = "y"
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    excel_file.name = "test.xlsx"
    # Test getting sheet names
    options = handler.get_options(excel_file)
    assert "sheet_name" in options
    assert "Sheet1" in options["sheet_name"]
    assert "Sheet2" in options["sheet_name"]
    # Test reading with sheet selection
    excel_file.seek(0)
    df = handler.read_data(excel_file, sheet_name="Sheet1")
    assert isinstance(df, pd.DataFrame)
    assert df.shape == (1, 2)
    assert list(df.columns) == ["a", "b"]


def test_json_handler():
    handler = JSONHandler()
    # Test file type detection
    assert handler.can_handle(MockFileName("test.json"))
    assert not handler.can_handle(MockFileName("test.csv"))

    # Test reading array-based JSON data
    json_data = {"a": [1, 3], "b": [2, 4]}
    file = BytesIO(json.dumps(json_data).encode())
    file.name = "test.json"
    df = handler.read_data(file)

    assert isinstance(df, pd.DataFrame)
    assert df.shape == (2, 2)
    assert list(df.columns) == ["a", "b"]
    assert df.iloc[0, 0] == 1
    assert df.iloc[1, 1] == 4

    # Test reading scalar JSON data (single object)
    scalar_json_data = {"fruit": "Apple", "size": "Large", "color": "Red"}
    file = BytesIO(json.dumps(scalar_json_data).encode())
    file.name = "test.json"
    df = handler.read_data(file)

    assert isinstance(df, pd.DataFrame)
    assert df.shape == (1, 3)
    assert set(df.columns) == {"fruit", "size", "color"}
    assert df.iloc[0]["fruit"] == "Apple"
    assert df.iloc[0]["size"] == "Large"
    assert df.iloc[0]["color"] == "Red"

    # Test reading list of scalar values
    scalar_list_json = [1, 2, 3, 4, 5]
    file = BytesIO(json.dumps(scalar_list_json).encode())
    file.name = "test.json"
    df = handler.read_data(file)

    assert isinstance(df, pd.DataFrame)
    assert df.shape == (5, 1)
    assert list(df.columns) == ["value"]
    assert df.iloc[0]["value"] == 1
    assert df.iloc[4]["value"] == 5

    # Test reading nested JSON data
    nested_json = {
        "person": {
            "name": "John",
            "age": 30,
            "address": {"city": "New York", "zip": "10001"},
        }
    }
    file = BytesIO(json.dumps(nested_json).encode())
    file.name = "test.json"
    df = handler.read_data(file)

    assert isinstance(df, pd.DataFrame)
    # Should be flattened into a dataframe with columns like "person.name", etc.
    assert "person.name" in df.columns or "person_name" in df.columns

    # Test options
    assert handler.get_options(None) == {}


def test_parquet_handler():
    # Skip test if pyarrow is not installed
    pytest.importorskip("pyarrow")

    handler = ParquetHandler()
    # Test file type detection
    assert handler.can_handle(MockFileName("test.parquet"))
    assert handler.can_handle(MockFileName("data.pq"))
    assert not handler.can_handle(MockFileName("test.csv"))

    # Create test DataFrame
    df_orig = pd.DataFrame({"a": [1, 3], "b": [2, 4]})

    # Save to BytesIO as parquet
    parquet_file = BytesIO()
    df_orig.to_parquet(parquet_file)
    parquet_file.seek(0)
    parquet_file.name = "test.parquet"

    # Test reading parquet data
    df = handler.read_data(parquet_file)

    assert isinstance(df, pd.DataFrame)
    assert df.shape == (2, 2)
    assert list(df.columns) == ["a", "b"]
    np.testing.assert_array_equal(df["a"].values, np.array([1, 3]))
    np.testing.assert_array_equal(df["b"].values, np.array([2, 4]))

    # Test options
    assert handler.get_options(None) == {}


def test_get_handler():
    # Test CSV handler
    csv_file = MockFileName("test.csv")
    handler = get_handler(csv_file)
    assert isinstance(handler, CSVHandler)

    # Test Excel handler
    excel_file = MockFileName("test.xlsx")
    handler = get_handler(excel_file)
    assert isinstance(handler, ExcelHandler)

    # Test JSON handler
    json_file = MockFileName("test.json")
    handler = get_handler(json_file)
    assert isinstance(handler, JSONHandler)

    # Test Parquet handler
    parquet_file = MockFileName("test.parquet")
    handler = get_handler(parquet_file)
    assert isinstance(handler, ParquetHandler)

    pq_file = MockFileName("data.pq")
    handler = get_handler(pq_file)
    assert isinstance(handler, ParquetHandler)

    # Test unsupported file type
    text_file = MockFileName("test.txt")
    handler = get_handler(text_file)
    assert handler is None
